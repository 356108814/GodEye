# encoding: utf-8
"""
excel工具类
@author Yuriseus
@create 2016-9-1 15:39
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import *
from openpyxl.writer.excel import ExcelWriter


class ExcelUtil(object):
    def __init__(self):
        pass

    @staticmethod
    def get_data_from_excel(excel_file, start_row=1):
        """
        从excel获取数据
        @param excel_file
        @param start_row
        @return list
        """
        data_list = []
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.worksheets[0]
        row_index = 1
        for row in ws.rows:
            if row_index >= start_row:
                data = []
                cell_index = 0
                for cell in row:
                    value = cell.value
                    data.append(value)
                    cell_index += 1
                data_list.append(data)
            row_index += 1
        return data_list

    @staticmethod
    def write_to_excel(excel_file, data_list, column_width=20.0):
        wb = Workbook()
        excel_writer = ExcelWriter(workbook=wb)
        sheet = wb.worksheets[0]
        sheet.name = ''
        merge_row = -1
        for row_index, row_list in enumerate(data_list):
            if len(row_list) == 1:    # 合并单元格
                merge_row = row_index + 1
            for cell_index, cell in enumerate(row_list):
                sheet.cell(row=row_index+1, column=cell_index+1).value = cell

        # 设置列宽度
        max_column = sheet.max_column
        for x in range(max_column):
            sheet.column_dimensions[chr(65+x)].width = column_width

        # 合并单元格，设置单元格样式
        if merge_row >= 0:
            sheet.merge_cells('A1:%s1' % (chr(65 + max_column - 1)), 1, 1, 1, max_column - 1)
            cell = sheet.cell('A1')
            font = Font(name=u'宋体', size=14, color=BLACK, bold=True)
            align = Alignment(horizontal='center', vertical='center')
            cell.font = font
            cell.alignment = align
        excel_writer.save(excel_file)

    @staticmethod
    def write_to_excel_by_sheet_data_list(excel_file, sheet_data_list):
        """
        数据写入多个sheet
        :param excel_file:
        :param sheet_data_list: sheet 数据字典列表
        :return:
        """
        wb = Workbook()
        excel_writer = ExcelWriter(workbook=wb)
        for sheet_index, sheet_data in enumerate(sheet_data_list):
            if sheet_index == 0:
                sheet = wb.worksheets[sheet_index]
            else:
                sheet = wb.create_sheet()
            data_list = sheet_data['data_list']
            column_width = 20.0
            merge_row = -1
            if 'title' in sheet_data:
                sheet.title = sheet_data['title']
            if 'column_width' in sheet_data:
                column_width = sheet_data['column_width']
            for row_index, row_list in enumerate(data_list):
                if len(row_list) == 1:    # 合并单元格
                    merge_row = row_index + 1
                for cell_index, cell_value in enumerate(row_list):
                    sheet.cell(row=row_index+1, column=cell_index+1).value = cell_value

            # 设置列宽度
            max_column = sheet.max_column
            for x in range(max_column):
                sheet.column_dimensions[chr(65+x)].width = column_width

            # 合并单元格，设置单元格样式
            if merge_row >= 0:
                sheet.merge_cells('A1:%s1' % (chr(65 + max_column - 1)), 1, 1, 1, max_column - 1)
                cell = sheet.cell('A1')
                font = Font(name=u'宋体', size=14, color=BLACK, bold=True)
                align = Alignment(horizontal='center', vertical='center')
                cell.font = font
                cell.alignment = align

        excel_writer.save(excel_file)
